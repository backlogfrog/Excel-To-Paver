#extract data from a spreadsheet and convert to Paver XML data for import

#finding files by extension in main dir/opening files, exec .py files
import os
import glob

#needed to parse date in scratchXml.py
from datetime import datetime

#colored text for funskies while I wait for data
import colorama
from colorama import Fore, Back, Style 
colorama.init()
#reset color/style after printing
colorama.init(autoreset=True)
#Colors, yay...
#Format: print(Fore.BLACK + "TEXT " + str(VARIABLE))
#Fore: BLACK, RED, GREEN, YELLOW, BLUE, MAGENTA, CYAN, WHITE, RESET.
#Back: BLACK, RED, GREEN, YELLOW, BLUE, MAGENTA, CYAN, WHITE, RESET.
#Style: DIM, NORMAL, BRIGHT, RESET_ALL

#openpyxl package opens excel sheets
from openpyxl import load_workbook

#NEW DATA HAS NO DATE - NEED TO SET DATE

##################################################################################
##################################################################################
##################################################################################
#set start positions to read data: 2012.xlsx data starts at row 4
#RowIncr=4
RowIncr=2   
##################################################################################
##################################################################################
##################################################################################
##################################################################################
##################################################################################




#import class info from inspectionClasses.py for inspection data and from mapping.py based on column to be printed to XML

from mapping import INSPECTED_SIZE, INSPECTED_DATE, INSPECTED_PID1, INSPECTED_PID2, DCOMMENT, P_LENGTH, P_WIDTH, SAMPLENUMBER, SWEATHERING_CODE, SWEATHERING_S, SWEATHERING_Q, ALLIGATOR_CODE, ALLIGATOR_S, ALLIGATOR_Q, BLOCKCRACK_CODE, BLOCKCRACK_S, BLOCKCRACK_Q, TRANSVERSE__CODE, TRANSVERSE_S, TRANSVERSE_Q, DEPRESSION_CODE, DEPRESSION_S, DEPRESSION_Q, POTHOLE_CODE, POTHOLE_S, POTHOLE_Q, EDGECRACKING_CODE, EDGECRACKING_S, EDGECRACKING_Q, JOINTSPALLING_CODE, JOINTSPALLING_S, JOINTSPALLING_Q, DURABILITYCRACKING_CODE, DURABILITYCRACKING_S, DURABILITYCRACKING_Q, FAULTING_CODE, FAULTING_S, FAULTING_Q, PATCHING_CODE, PATCHING_S, PATCHING_Q, BUMPSAG_CODE, BUMPSAG_S, BUMPSAG_Q, SAMPLESIZE

#set header/schema for xml
xml_header = "<?xml version=\"1.0\" encoding=\"utf-8\" ?>"
xml_schema = "<pavementData xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xsi:noNamespaceSchemaLocation=\"PavementInspectionDataV2.xsd\">"

#unneeded at this time
#needed to add date/time to JSON dictionary, 
#def json_serial(obj):
#    """JSON serializer for objects not serializable by default json code"""
#    if isinstance(obj, (datetime)):
#       return obj.isoformat()
#    raise TypeError ("Type %s not serializable" % type(obj))

#############################open workbook from xlsx files in main dir for mult choice
#############################

excel_list = [f for f in glob.glob("*.xlsx")]
print(Fore.BLUE + "Available Spreadsheets:")
for i, db_name in enumerate(excel_list, start=1):
    print ('{}. {}'.format(i, db_name))
  
while True:
	try:
		selected = int(input(Fore.CYAN + 'Select spreadsheet (1-{}): '.format(i)))
		db_name = excel_list[selected-1]
		print(Style.DIM + 'Loading {}'.format(db_name), "\n")
		break
	except (ValueError, IndexError):
		print(Fore.BLUE + 'Invalid. Please enter number between 1 and {}!'.format(i))
		print("	")

#load input/workbook from above and set current active sheet to "sheet"
workbook = load_workbook(filename=db_name, data_only=True)
workbook.sheetnames
sheet = workbook.active


#set/show last row and column to avoid out of range error
#LastRow=sheet.max_row
#LastColumn=sheet.max_column
#print ("Last Column ", LastColumn)


#something is fucky here, last row shows is greater than 
LastRow=sheet.max_row-RowIncr
print("Total row number is ", LastRow)


########################################## - unwanted as of 2.19.20 - commented out just auto pulls all rows, no longer needed for debugging
#request user input to determine how many records are wanted, keep within range of data
#based on the LastRow, number is off by 1, but it works at this time
#while True:
#    number = input(Fore.CYAN + "How many rows? ")
#    try:
#        number = int(number)
#    except ValueError:
#        print(Fore.BLUE + number, 'is not a number, try again.')
#        continue
#    if number <= 0:
#        print("    ")
#        print(Fore.BLUE + "Too low, try again.")
#    elif number > LastRow:
#        print("    ")
#        print(Fore.BLUE + "Too high, try again.")
#    else:
#        print("    ")
#        break
#set row input to final row read
#LastRow=number
##########################################################################




#LastRow = Copy of 2012.xlsx 685 readable lines
#2012.xlsx has 686 readable lines
FinalRow = sheet.max_row

#check to see if last row has data for first column - if 0 then subtract 1 from last Row
#to avoid empty data/date error
#else, add LastRow to RowIncr(sheet actual-data-start offset)

#DETERMINE IF FURTHER ROWS ARE BLANK/0 AND stop pulling data, im not sure how this works currently 2.19.20
cellCheck=sheet.cell(row=FinalRow, column=1).value
if cellCheck == 0:
	LastRow=LastRow+RowIncr-1
else:
	LastRow=LastRow+RowIncr


#set filename for output and log based on EXCEL name
#request filename for output
#no longer needed for debugging/etc 2.19.20
#exec(open("inputStart.py").read())

#cut filename of .xml/.log file to 4 char
fileName=db_name[0:4]

#write files/delete if present
if os.path.exists(fileName+".xml"):
  os.remove(fileName+".xml")
  os.remove(fileName+".log")
  print(Style.DIM + Fore.RED +"Original File Deleted \n \n")
else:
  print(Style.DIM + Fore.BLUE +"New file created \n \n")
f = open(fileName + ".xml", "a+")
logFile = open(fileName + ".log", "a+")






#add headers to top of page
print(xml_header, "\n", xml_schema, sep="", file=f)

#how many rows read/written
rowsRead=0
ticker = 0

def emptyData():
	print ("Empty: ", row[INSPECTED_PID2],":", row[SAMPLENUMBER], " ###################", file=logFile)
	print ("Empty: ", row[INSPECTED_PID2],":", row[SAMPLENUMBER], " ###################")

def fullData():
		print("Data: ", row[INSPECTED_PID2], ":", row[SAMPLENUMBER], file=logFile)
		print("Data: ", row[INSPECTED_PID2], ":", row[SAMPLENUMBER])

#check each code to ensure there is a distress for this row, set ticker to 1 to write
def codeCheck(code):
	
	try:
		if float(code) > 0:
			global ticker
			ticker = 1
			fullData()
			#print(row[INSPECTED_PID2])
		else:
			emptyData()
			
	except ValueError:
		#print ("empty")
		#print(row[INSPECTED_PID2])
		emptyData()
		ticker = 0

#TESTING for specific errors in specific PID import
#PIDREQUEST = input("Enter your PID for PID specific xml: ") 
#print("PID: ", PIDREQUEST) 
#for if statement below: if ticker == 1 and row[INSPECTED_PID2] == int(PIDREQUEST)


for row in sheet.iter_rows(min_row=RowIncr, max_row=LastRow, values_only=True):
			rowsRead=rowsRead+1
			#iteration check through a dictionary/list wasn't fucking working, so as a "temp" (perm) fix, a function check to force float actually worked instead of type errors for ">""
			ticker = 0
			codeCheck(row[SWEATHERING_CODE])	 	
			codeCheck(row[ALLIGATOR_CODE])
			codeCheck(row[BLOCKCRACK_CODE])
			codeCheck(row[TRANSVERSE__CODE])
			codeCheck(row[DEPRESSION_CODE])
			codeCheck(row[POTHOLE_CODE])
			codeCheck(row[EDGECRACKING_CODE])
			codeCheck(row[JOINTSPALLING_CODE])
			codeCheck(row[DURABILITYCRACKING_CODE])
			codeCheck(row[FAULTING_CODE])
			codeCheck(row[PATCHING_CODE])
			codeCheck(row[BUMPSAG_CODE])
			#write data to xml if the code is actually greater than 0
			if ticker == 1:
				#print("PIDREQUEST")
				exec(open("scratchXml.py").read())
			
			

#print (Fore.RED + "FIRST SS IS IN 2010 - NO SET DATE \n \n")
	
#close xml main tag and workbook
print("</pavementData>", sep="", file=f)
f.close()
logFile.close()

#print rows and files read
print(Fore.MAGENTA + "\nRows Read: ", int(rowsRead)-1)		
print(Fore.MAGENTA + "File Opened:", db_name)